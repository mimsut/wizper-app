# -*- coding: utf-8 -*-
import json, openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# (key, ko, change, note)  change: 신규/분리/변수명수정/문구수정/유지
CAT = [
 # ── 온보딩 · 로그인 (Google 계정 전용) ──
 # 인증은 Google 계정으로만. 휴대폰 인증 없음. 성별·연락처 등 기본 정보와 연구 ID는
 # 관리자가 어드민에서 이메일 기준으로 등록·발급 (앱에서는 입력받지 않음).
 ("onb_account_title","위즈퍼링에 로그인","문구수정","계정 생성 폼 → Google 로그인"),
 ("onb_account_sub","연구 담당자가 등록한 Google 계정으로 로그인해 주세요","문구수정","안내문 변경"),
 ("onb_login_google","Google로 계속하기","신규","유일한 로그인 수단"),
 ("onb_login_note","휴대폰 인증 없이 Google 계정으로만 로그인해요. 성별·연락처 등 기본 정보는 연구 담당자가 등록해요.","신규","로그인 정책 안내"),
 ("onb_login_signed","로그인됐어요","신규",""),
 ("onb_login_issued_id","발급된 연구 ID","신규","이메일 기준 관리자 발급"),
 ("onb_login_info","관리자가 등록한 기본 정보","신규",""),
 ("common_next","다음","유지",""),
 # ── 온보딩 · 페어링 ──
 ("onb_pairing_title","WIZPR RING을 연결해 주세요","문구수정","제목 변경, 끝 개행 제거"),
 ("onb_pairing_sub","반지를 손가락에 착용한 상태로 진행해 주세요","유지",""),
 ("onb_pairing_scanning","연결 가능한 WIZPR RING을 검색 중입니다.\n잠시만 기다려 주세요.","문구수정","안내문 변경 (2줄 의도된 개행)"),
 ("onb_pairing_connect","연결","문구수정","'연결하기' → '연결'"),
 ("onb_pairing_rescan","다시 검색","유지",""),
 ("onb_pairing_help","페어링에 문제가 있나요?","유지",""),
 # ── 온보딩 · 권한 (라벨/설명 분리, em-dash 제거) ──
 ("onb_perm_title","연구에 필요한 권한을\n허용해 주세요","유지",""),
 ("onb_perm_sub","권한이 꺼져 있는 동안의 기록은 다시 수집할 수 없어요","유지",""),
 ("onb_perm_noti","알림","분리","'알림 — 설명' em-dash 라벨/설명 분리"),
 ("onb_perm_noti_desc","설문 시간을 알려드려요","분리",""),
 ("onb_perm_loc","위치","분리","em-dash 분리"),
 ("onb_perm_loc_desc","외출 반경과 이동 거리를 계산해요","분리",""),
 ("onb_perm_activity","신체활동","분리","em-dash 분리"),
 ("onb_perm_activity_desc","걸음 수와 활동 시간을 수집해요","분리",""),
 ("onb_perm_bt","Bluetooth","분리","em-dash 분리"),
 ("onb_perm_bt_desc","반지와 연결을 유지해요","분리",""),
 ("onb_perm_battery","배터리 사용 제한 해제","분리","em-dash 분리"),
 ("onb_perm_battery_desc","백그라운드 수집이 끊기지 않아요","분리",""),
 ("onb_perm_go_settings","설정하러 가기","유지",""),
 ("onb_perm_allowed","허용됨","유지",""),
 ("onb_perm_privacy","위치 좌표나 통화 내용 원문은 서버로 보내지 않아요. 기기 안에서 계산한 값만 전송돼요.","유지",""),
 # ── 온보딩 · 시간 설정 ──
 ("onb_time_title","첫 설문 시간을\n골라 주세요","유지",""),
 ("onb_time_sub","하루 4번, 4시간 간격으로 진행돼요. 일과를 시작한 뒤의 시간을 고르면 참여하기 좋아요.","유지",""),
 ("onb_time_manual","직접 선택","유지",""),
 ("onb_time_auto","자동 설정","유지",""),
 ("onb_time_lock","연구 시작 후에는 시간을 바꿀 수 없어요. 각 회차는 알림 후 60분 동안 참여할 수 있어요.","유지",""),
 ("onb_time_confirm_title","이 시간으로 확정할까요?","유지",""),
 ("onb_time_confirm_body","설정한 시간으로 설문 알림이 발송돼요.\n연구가 시작되면 직접 변경할 수 없어요.","유지",""),
 ("onb_time_confirm_yes","확정하기","유지",""),
 ("onb_time_confirm_no","다시 고르기","유지",""),
 ("onb_ready_title","준비가 끝났어요","유지",""),
 ("onb_ready_first_ema","첫 설문 예정","유지",""),
 ("onb_ready_start","앱 시작하기","유지",""),
 # ── 홈 ──
 ("home_greeting","안녕하세요, {id}님","유지","{id}=연구 ID"),
 ("home_greeting_sub","오늘도 소중한 참여 부탁드려요","유지",""),
 ("home_today_earned","오늘 적립액","유지",""),
 ("home_total","총 누적액","유지",""),
 ("home_reset_note","매일 자정에 총 누적액으로 옮겨져요","유지",""),
 ("home_bonus_banner","오늘 모든 참여를 완료하면 보너스 800원을 드려요","문구수정","일일 완료 보너스 1,000 → 800원 (보상 표 기준)"),
 ("home_today_survey","오늘의 설문","유지",""),
 ("home_survey_sub","하루 4번, 각 회차 1시간 이내에 참여할 수 있어요","문구수정","60분 → 1시간"),
 ("home_join","참여하기","유지",""),
 ("home_done","완료","유지",""),
 ("home_opens_at","{n}시에 열려요","유지","{n}=회차 시각(시)"),
 ("home_today_voice","오늘의 음성 과제","유지",""),
 ("home_progress","오늘의 진행","유지",""),
 ("home_calendar","참여 현황","유지",""),
 # ── EMA 설문 ──
 ("ema_title","설문 응답","문구수정","'{t} 자기보고 설문' → '설문 응답' ({t} 제거)"),
 # 외로움
 ("ema_loneliness_instruction","지난 설문 시점으로부터 현재까지 다음의 문제에 해당하는 정도를 선택해 주세요.","신규","외로움 척도 안내"),
 ("ema_loneliness_ask_1","나는 나와 같이 있어 줄 사람이 부족하다.","신규",""),
 ("ema_loneliness_ask_2","나는 혼자 남겨진 것 같다고 느껴진다.","신규",""),
 ("ema_loneliness_ask_3","나는 사람들 사이에서 고립되어 있다고 느껴진다.","신규",""),
 ("ema_loneliness_scale_1","전혀 그렇지 않다","신규",""),
 ("ema_loneliness_scale_2","가끔 그렇다","신규",""),
 ("ema_loneliness_scale_3","자주 그렇다","신규",""),
 # PHQ-9
 ("ema_phq_instruction","지난 설문 시점으로부터 현재까지 다음의 문제에 해당하는 정도를 1~5까지 선택해 주세요.","신규","PHQ-9 안내"),
 ("ema_phq_ask_1","기분이 가라앉거나, 우울하거나, 희망이 없다고 느낌","신규",""),
 ("ema_phq_ask_2","평소 하던 일에 대한 흥미가 없어지거나 즐거움을 느끼지 못함","신규",""),
 ("ema_phq_ask_3","잠들기가 어렵거나 자주 깸, 혹은 너무 많이 잤음","신규",""),
 ("ema_phq_ask_4","평소보다 식욕이 줄었음, 혹은 평소보다 많이 먹음","신규",""),
 ("ema_phq_ask_5","다른 사람들이 눈치챌 정도로 평소보다 말과 행동이 느려짐, 혹은 너무 안절부절못해서 가만히 앉아 있을 수 없음","신규",""),
 ("ema_phq_ask_6","피곤하거나 기운이 없음","신규",""),
 ("ema_phq_ask_7","내가 잘못했거나 실패했다는 생각이 듦, 혹은 자신과 가족을 실망시켰다고 생각함","신규",""),
 ("ema_phq_ask_8","신문을 읽거나 TV를 보는 것과 같은 일상적인 일에도 집중할 수 없었음","신규",""),
 ("ema_phq_ask_9","차라리 죽는 것이 더 낫겠다고 생각함, 혹은 자해할 생각을 함","신규",""),
 ("ema_phq_scale_1","전혀 해당되지 않음","신규","척도 최저 (1)"),
 ("ema_phq_scale_5","매우 해당함","신규","척도 최고 (5)"),
 # GAD-7
 ("ema_gad_instruction","지난 설문 시점으로부터 현재까지 다음의 문제에 해당하는 정도를 1~5까지 선택해 주세요.","신규","GAD-7 안내"),
 ("ema_gad_ask_1","초조하거나 불안하거나 조마조마하게 느낀다.","신규",""),
 ("ema_gad_ask_2","걱정하는 것을 멈추거나 조절할 수가 없다.","신규",""),
 ("ema_gad_ask_3","여러 가지 것들에 대해 걱정을 너무 많이 한다.","신규",""),
 ("ema_gad_ask_4","편하게 있기가 어렵다.","신규",""),
 ("ema_gad_ask_5","너무 안절부절못해서 가만히 있기가 힘들다.","신규",""),
 ("ema_gad_ask_6","쉽게 짜증이 나거나 쉽게 성을 내게 된다.","신규",""),
 ("ema_gad_ask_7","마치 끔찍한 일이 생길 것처럼 두렵게 느껴진다.","신규",""),
 ("ema_gad_scale_1","전혀 해당되지 않음","변수명수정","'ema_gad_1' → 'ema_gad_scale_1'"),
 ("ema_gad_scale_5","매우 해당함","변수명수정","'ema_gad_5' → 'ema_gad_scale_5'"),
 # ESM
 ("ema_esm_instruction","알림이 울리기 직전의 상태를 떠올리고 답해 주세요.\n각 문항에 대해 1(전혀 아니다) ~ 7(매우 그렇다) 중 해당하는 정도를 선택해 주세요.","신규","ESM 안내"),
 ("ema_esm_ask_1","내 생각을 말로 표현하기 어렵다.","신규",""),
 ("ema_esm_ask_2","내 생각이 다른 사람의 영향을 받는다.","신규",""),
 ("ema_esm_ask_3","어떤 생각을 머릿속에서 떨쳐낼 수 없다.","신규",""),
 ("ema_esm_ask_4","내가 현실이 아닌 것처럼 느껴진다.","신규",""),
 ("ema_esm_ask_5","누군가 나를 해치려 한다는 의심이 든다.","신규",""),
 ("ema_esm_ask_6","실제로는 없는 소리가 들린다.","신규",""),
 ("ema_esm_ask_7","실제로는 없는 것이 보인다.","신규",""),
 ("ema_esm_ask_8","내가 통제력을 잃어가고 있는 것 같다.","신규",""),
 ("ema_esm_scale_1","전혀 아니다","신규","척도 최저 (1)"),
 ("ema_esm_scale_7","매우 그렇다","신규","척도 최고 (7)"),
 ("ema_next_blocked","모든 문항에 답하면 넘어갈 수 있어요","유지",""),
 ("ema_exit_title","설문을 그만할까요?","유지",""),
 ("ema_exit_body","지금 그만두면 응답이 저장되지 않고\n이번 회차 보상 250원도 받을 수 없어요.","유지",""),
 ("ema_exit_continue","계속 응답하기","유지",""),
 ("ema_exit_quit","그만하기","유지",""),
 ("ema_done_title","250원이 적립됐어요","유지",""),
 ("ema_done_sub","오늘 {n}번째 설문까지 완료했어요","유지","{n}=오늘 완료 회차 수"),
 ("ema_done_next","다음 과제는 {t}시에 열려요. 알림으로 알려드릴게요.","문구수정","'설문'→'과제', {t} 뒤 '시' 추가 ({t}=시각 숫자)"),
 ("common_home","홈으로","유지",""),
 # ── 음성 과제 ──
 ("voice_title","음성 대화 과제","유지",""),
 ("voice_topic_label","오늘의 대화 주제","유지",""),
 ("voice_topic_sub","답변에 따라 이어지는 질문을 드려요","유지",""),
 ("voice_guide_quiet","조용한 곳에서 편하게 이야기해 주세요","유지",""),
 ("voice_guide_time","최소 20초, 최대 2분 동안 진행돼요","유지",""),
 ("voice_guide_free","정답은 없어요. 떠오르는 대로 말해 주세요.","문구수정","마침표 추가"),
 ("voice_start","대화 시작하기","유지",""),
 ("voice_start_reward","완료하면 700원이 적립돼요","문구수정","500원 → 700원"),
 ("voice_ring_connected","WIZPR RING 연결됨","유지",""),
 ("voice_ring_required","반지를 연결해야 시작할 수 있어요","유지",""),
 ("voice_followup_note","이어지는 질문 · 음성으로도 들려드려요","유지",""),
 ("voice_min_ok","최소 시간을 채웠어요. 지금 마쳐도 돼요","문구수정","em-dash 제거"),
 ("voice_pause_hint","말이 3초 이상 끊기면 안내를 드려요","유지",""),
 ("voice_finish","녹음 완료","유지",""),
 ("voice_done_title","700원이 적립됐어요","문구수정","500원 → 700원"),
 ("voice_done_sub","오늘 과제를 모두 마쳤어요","유지",""),
 ("voice_bonus_title","모든 참여 완료 보너스","유지",""),
 ("voice_bonus_sub","설문 4회 + 대화 1회","유지",""),
 ("voice_next_day","다음 참여일","유지",""),
 # ── 기기 · 설정 ──
 ("device_title","기기 연결","유지",""),
 ("device_connected","연결됨 · 방금 동기화","유지",""),
 ("device_battery","반지 배터리","유지",""),
 ("device_last_sync","마지막 데이터 전송","유지",""),
 ("device_auto_reconnect","연결이 끊기면 자동으로 다시 연결을 시도해요. 계속 실패하면 알림으로 알려드려요.","유지",""),
 ("device_reconnect","다시 연결하기","유지",""),
 ("settings_title","설정","유지",""),
 ("settings_doc","연구 설명문 보기","유지",""),
 ("settings_perm","권한 상태 확인","유지",""),
 ("settings_perm_alert","{n}개 확인 필요","유지","{n}=미허용 권한 수"),
 ("settings_noti","알림 설정","유지",""),
 ("settings_ema_time","설문 시간","유지",""),
 ("settings_ema_locked","{t} 시작 · 변경 불가","유지","{t}=1회차 시각"),
 ("settings_contact","연구팀에 문의하기","유지",""),
 ("settings_withdraw","연구 참여 철회","유지",""),
 ("settings_withdraw_note","참여 철회 시 이후 수집이 즉시 중단돼요. 자세한 내용은 연구 설명문을 확인해 주세요.","유지",""),
 # ── 푸시 알림 ──
 ("push_session_open","{t} 설문에 참여할 수 있어요. 60분 동안 열려 있어요.","유지","{t}=회차 시각"),
 ("push_day_start","오늘은 설문 및 대화과제 참여가 가능한 날입니다. 오늘 4번의 설문과 1번의 대화과제를 모두 참여하시면 2,500원이 지급됩니다.","문구수정","1,500원 → 2,500원"),
 ("push_ema_low","오늘은 설문 및 대화과제 참여가 가능한 날입니다. 지난 8회의 설문 중 총 {n}회 참여하셨습니다. 원활한 연구 진행을 위해 성실히 참여해주시면 감사하겠습니다!","유지","{n}=참여 횟수"),
 ("push_sensor_low","어제의 센서 수집이 원활하지 않았습니다. 다음 안내에 따라 조치하신 후 설문에 참여해 주세요.","유지",""),
 ("push_voice_low","최근 대화형 과제에 {n}회 연속 참여하지 않으셨습니다. 잊지 말고 참여해 주세요!","유지","{n}=연속 미응답 횟수"),
 ("push_reminder","이번 회차 EMA 응답 가능 시간이 30분밖에 남지 않았습니다. 응답 시간이 지나면 이번 회차는 참여가 어려우니, 잊지 마시고 꼭 확인 부탁드립니다.","유지",""),
]

# ---- write corrected xlsx ----
wb=openpyxl.Workbook(); ws=wb.active; ws.title="다국어 시트"
hdr=["key","ko","변경유형","비고"]
ws.append(hdr)
for c in ws[1]:
    c.font=Font(bold=True); c.fill=PatternFill("solid",fgColor="EEEDFC")
for key,ko,chg,note in CAT:
    ws.append([key,ko,chg,note])
for row in ws.iter_rows(min_row=2):
    row[1].alignment=Alignment(wrap_text=True, vertical="top")
    row[3].alignment=Alignment(wrap_text=True, vertical="top")
ws.column_dimensions["A"].width=28
ws.column_dimensions["B"].width=60
ws.column_dimensions["C"].width=12
ws.column_dimensions["D"].width=40
out="/Users/user/Downloads/위즈퍼링_카피라이팅_시트_v2_반영.xlsx"
wb.save(out)

# ---- write copy.js ----
ko_map={k:v for k,v,_,_ in CAT}
js="// 위즈퍼링 카피 (ko) — 카피라이팅 시트 v2 반영 자동 생성. 편집은 시트에서.\n"
js+="export const KO = "+json.dumps(ko_map, ensure_ascii=False, indent=2)+";\n\n"
js+="export function t(key, vars) {\n  let s = KO[key] || key;\n  if (vars) for (const k in vars) s = s.split('{'+k+'}').join(vars[k]);\n  return s;\n}\n"
open("/Users/user/wizper-app/src/copy.js","w").write(js)

# summary counts
from collections import Counter
cnt=Counter(c for _,_,c,_ in CAT)
print("sheet:", out)
print("rows:", len(CAT), "| change types:", dict(cnt))
print("copy.js written:", "/Users/user/wizper-app/src/copy.js")
